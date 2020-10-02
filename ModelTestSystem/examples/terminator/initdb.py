from synbiomts import dbms
from Bio import SeqIO
import xlrd
from openpyxl import load_workbook
get = lambda cell: cell[0].value # for openpyxl

# Initialize DataBase
DB = dbms.DataBase()

'''
-----------------------------------------------------------------------------------
Cambray, Guillaume, Joao C. Guimaraes, Vivek K. Mutalik,Colin Lam,
Quynh-Anh Mai, Tim Thimmaiah, James M. Carothers, Adam P. Arkin, and Drew Endy.
"Measurement and modeling of intrinsic transcription terminators."
Nucleic acids research 41, no. 9 (2013): 5139-5148.'''

# test vector on Addgene: http://www.addgene.org/47846/
pFAB763 = SeqIO.read('datasets/pFAB763-trp-L126.gb','genbank')

for f in pFAB763.features:
    label = f.qualifiers['label'][0]
    if label == 'PLtetO-1 promoter':
        start = f.location.end # TSS for mRNA transcript
    elif label == 'GoldenGate-tgat':
        i = f.location.end # insertion site for terminator
    elif label == 'GoldenGate-ggcg':
        j = f.location.start # insertion site for terminator
    elif label == 'rrnB T1 terminator':        
        end = f.location.end # transcript end of bicistronic operon

operon = str(pFAB763.seq[start:end])
i -= start
j -= start

# supplementary information
wb = xlrd.open_workbook('datasets/Cambray_2013.xls')
sheet = wb.sheet_by_index(1)

dataset = {
    'NAME': sheet.col_values(colx=1, start_rowx=4, end_rowx=58),
    'SOURCE': sheet.col_values(colx=3, start_rowx=4, end_rowx=58),
    'COORDINATES': sheet.col_values(colx=4, start_rowx=4, end_rowx=58),
    'TERMINATOR.SEQ': sheet.col_values(colx=6, start_rowx=4, end_rowx=58),
    'EFFICIENCY': sheet.col_values(colx=8, start_rowx=4, end_rowx=58),
    'ORGANISM': 'Escherichia coli BW25113',
    'TEMP': 37.0,
    'DATASET': 'Cambray_2013'
}

terminators = dataset['TERMINATOR.SEQ']
dataset['OPERON'] = [operon[:i]+t+operon[j:] for t in terminators]
dataset['TERMINATOR.START'] = i
dataset['TERMINATOR.END'] = [i+len(t)-1 for t in terminators]

DB += dataset # add dataset to DataBase


'''
-----------------------------------------------------------------------------------
Ying-Ja Chen, Peng Liu, Alec A K Nielsen, Jennifer A N Brophy,
Kevin Clancy, Todd Peterson & Christopher A Voigt
"Characterization of 582 natural and synthetic terminators andquantification of their design constraints"
Nature Methods, 2013, Vol. 10, No. 7; doi:10.1038/nmeth.2515'''

# relationship defined between terminator strength (TS) and efficiency (TE)
TE = lambda TS: 1-1/TS

# Natural terminators (Supplementary Table S2)
wb = load_workbook(filename='datasets/Chen_S2_2013.xlsx',read_only=True)
ws = wb['Fig2a-natural']

dataset = {
    'NAME': map(get,ws['A2':'A318']),
    'TERMINATOR.SEQ': map(get,ws['F2':'F318']),
    'EFFICIENCY': map(lambda TS: 1-1/TS, map(get,ws['L2':'L318'])),
    'SOURCE': "NC_000913",
    'ORGANISM': 'Escherichia coli DH5-alpha',
    'TEMP': 37.0,
    'DATASET': 'Chen_2013'
}

# Synthetic terminators (Supplementary Table S3)
wb = load_workbook(filename='datasets/Chen_S3_2013.xlsx',read_only=True)
ws = wb['Sheet1']

dataset['NAME'] += map(get,ws['A2':'A266'])
dataset['TERMINATOR.SEQ'] += map(get,ws['E2':'E266'])
dataset['EFFICIENCY'] += map(TE, map(get,ws['K2':'K266']))

# test vector on Addgene: http://www.addgene.org/46002/
pGR = SeqIO.read('datasets/pGR.gb','genbank')

for f in pGR.features:
    label = f.qualifiers['label'][0]
    if label == 'araBAD promoter':
        start = f.location.end # TSS for mRNA transcript
    elif label == 'EcoRI':
        i = f.location.end # insertion site for terminator
    elif label == 'SpeI':
        j = f.location.start # insertion site for terminator
    elif label == 'rrnB T1 terminator':        
        end = f.location.end # transcript end of bicistronic operon

operon = str(pGR.seq[start:end])
i -= start
j -= start

terminators = dataset['TERMINATOR.SEQ']
dataset['OPERON'] = [operon[:i]+t+operon[j:] for t in terminators]
dataset['TERMINATOR.START'] = i
dataset['TERMINATOR.END'] = [i+len(t)-1 for t in terminators]

DB += dataset

DB.save('terminators',type='pickle')